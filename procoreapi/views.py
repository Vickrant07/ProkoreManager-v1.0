from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
import requests
from .models import *
from apscheduler.schedulers.background import BackgroundScheduler
import openpyxl
from django.http import HttpResponse

######################################################################################################
# Procore api connection functions for getting data from Procore and storing it in the database
######################################################################################################

# Procore Live credentials
OAUTH_URL = 'https://login.procore.com'
BASE_URL = 'https://api.procore.com'
CLIENT_ID = '2HQPliajY5fTKDSaJKtVb1UhghT5nnDm0FSzY31aTbU'
CLIENT_SECRET = 'JSsjOeDjSvHAGIbe7Cr-PTSV_qOJFlEcML3K9jcxhGU'
REDIRECT_URI = 'urn:ietf:wg:oauth:2.0:oob'

# Procore Monthly sandbox credentials
# OAUTH_URL = 'https://login-sandbox-monthly.procore.com'
# BASE_URL = 'https://api-monthly.procore.com'
# CLIENT_ID = '78226445873eeed7ac57715df48830f52b6ce6a2fba1c31099a8afd67e26a5d6'
# CLIENT_SECRET = 'c46408666b71806645b0d686281c64148b68dec8fd07980815d357d756d87710'
# REDIRECT_URI = 'urn:ietf:wg:oauth:2.0:oob'

# Procore sandbox credentials
# OAUTH_URL = 'https://login-sandbox.procore.com'
# BASE_URL = 'https://sandbox.procore.com'
# REDIRECT_URI = 'urn:ietf:wg:oauth:2.0:oob'
# CLIENT_ID = 'e6194919d35cdde9440065c0d95468b058a273a51f2ecf6ac99b60b283bc7c01'
# CLIENT_SECRET = 'b98e1bc22386a63ac3e0ff5c4cb569c39256314c0efe1a1554418d804b9ff48b'

######################################################################################################
# Functions for getting credentials from Procore
######################################################################################################
# function to get fresh access token from Procore whenever needed

def get_acess_token():
    '''
    function to get the access token from Procore for 'client credentials' Auth type.
    '''
    post_data = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "grant_type": "client_credentials",
        "redirect_uri": REDIRECT_URI
    }
    response = requests.post(OAUTH_URL+"/oauth/token", data=post_data)
    response_json = response.json()
    # created_at = response_json['created_at']
    # expires_in = response_json['expires_in']
    access_token = response_json['access_token']
    # print(access_token)
    return access_token

# Function to get our company id based on access token whenever needed
def get_company_id():
    '''
    function to get the company id from Procore for API requests. 
    '''
    access_token = get_acess_token()
    url = BASE_URL + "/rest/v1.0/companies"
    headers = {
        "Authorization": "Bearer " + access_token
    }
    response = requests.get(url, headers=headers)
    response_json = response.json()                
    # print(response_json)
    company_id = response_json[0]['id']
    # print(company_id)
    return company_id

### Get Project pm for a project based on id
def get_project_manager_for_project(project_id, company_id):
    '''Get pm'''
    access_token = get_acess_token()
    company_id = company_id
    url = BASE_URL + '/rest/v1.0/project_roles'
    headers = {
        "Authorization": "Bearer " + access_token,
        "Procore-Company-Id": str(company_id)
    }
    payload = {
        'project_id' : project_id,
    }
    response = requests.get(url, headers = headers, json = payload)
    response_json = response.json()
    # print(response_json)   
    if len(response_json) > 0:
        for item in response_json:
            if item["role"] == 'Project Manager':
                project_manager_raw, ignore = item["name"].split("(")
                project_manager = project_manager_raw.strip()
                # print(project_manager)
                pms_in_db = ProjectManagers.objects.values_list('project_manager', flat=True)
                pms_in_db_list = list(pms_in_db)
                # print(pms_in_db_list)
                if project_manager not in pms_in_db_list:
                    projectManager = ProjectManagers()
                    projectManager.project_manager = project_manager
                    projectManager.save()
                return project_manager
    else:
        project_manager = "N/A"
        return project_manager 

# Get all WHS projects based on access token and company id and save to database

# EWI = custom_field_64757
# CWI = custom_field_64762
# IWI = custom_field_64763
# Window = custom_field_64764
# Plumbing = custom_field_64765
# Attic = custom_field_64766
# RTV = custom_field_64767
# MEV = custom_field_64768
# Draughtproofing = custom_field_64769
# ATT = custom_field_64771
# Solar = custom_field_66383
# Floor = custom_field_67248
# MVHR = custom_field_67249
# DCV = custom_field_67250
# QC_Mid_Final = custom_field_67252
# C/O Invoices = custom_field_82331
# Final BER = custom_field_82333
# DOW = custom_field_562949953936775

#####################################################################################################
# Functions to run on regular intervals for keeping database uptodate with procore
#####################################################################################################

#Run once a day: Get all current active WHS projects and save to database
def get_all_active_OSS_and_NISEP_projects():
    '''
    Get all the WHS projects in the Procore and save to the database.
    '''
    access_token = get_acess_token()
    company_id = get_company_id()
    page = 0
    boolean = True
    count = 0
    
    while boolean == True:
        url = BASE_URL + '/rest/v1.1/projects?company_id='+str(company_id)+'&page='+str(page)+'&per_page=20'
        headers = {
            "Authorization": "Bearer " + access_token
            
        }
        response = requests.get(url, headers=headers)
        response_json = response.json()
        # print(response_json)
        # print(len(response_json))

        # ### Test code to find out which custom fields are which, based on custom field numbers
        # project_id = response_json['id']
        # for item in response_json['custom_fields']:
        #     print (item)
        #     custom_field_id = item.split('_')[2]
        #     print(custom_field_id)
        #     url = BASE_URL + '/rest/v1.0/custom_field_definitions/'+str(custom_field_id)
        #     headers = {
        #         "Authorization": "Bearer " + access_token,
        #         "Procore-Company-Id": str(company_id)
        #     }
        #     payload = {
        #         'project_id' : project_id,
        #         'company_id' : str(company_id),
        #     }
        #     response = requests.get(url, headers=headers, json=payload)
        #     print(response)
        #     response_json = response.json()
        #     print(response_json)
        
        if len(response_json) > 0:
            
            for item in response_json:
                project = ProjectDetails()
                # print(item['project_id'])
                
                if item['project_number'] != None:
                    striped_project_nmumber = item['project_number'].strip()
                    # print(striped_project_nmumber)

                    if (item['project_stage']['name'][0] == 'O' and item['project_stage']['name'][1] == 'S') or (item['project_stage']['name'][0] == 'N' and item['project_stage']['name'][1] == 'I') :
                        count+=1
                        # print(striped_project_nmumber)
                        project_id = item['id']
                        project.project_id = project_id
                        project.project_number = striped_project_nmumber
                        project.project_name = item['name']
                        if item['project_stage']['name'][0] == 'O':
                            project.project_office = 'OSS'
                            if 'OSS' not in ProjectOffices.objects.values_list('project_offices', flat=True):
                                project_office = ProjectOffices()
                                project_office.project_offices = 'OSS'
                                project_office.save()
                        else:
                            project.project_office = 'NISEP'
                            if 'NISEP' not in ProjectOffices.objects.values_list('project_offices', flat=True):
                                project_office = ProjectOffices()
                                project_office.project_offices = 'NISEP'
                                project_office.save()
                        project.project_stage = item['project_stage']['name']
                        
                        if item['custom_fields']['custom_field_64757']['value'] != None:
                            ewi_stage = item['custom_fields']['custom_field_64757']['value']['label']
                            project.EWI_STATUS = ewi_stage
                        # else:
                        #    .EWI_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_64762']['value'] != None:
                            cwi_stage = item['custom_fields']['custom_field_64762']['value']['label']
                            project.CWI_STATUS = cwi_stage
                        # else:
                        #    .CWI_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_64763']['value'] != None:
                            iwi_stage = item['custom_fields']['custom_field_64763']['value']['label']
                            project.IWI_STATUS = iwi_stage
                        # else:
                        #    .IWI_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_64764']['value'] != None:
                            window_stage = item['custom_fields']['custom_field_64764']['value']['label']
                            project.WINDOW_STATUS = window_stage
                        # else:
                        #    .WINDOW_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_64765']['value'] != None:
                            plumbing_stage = item['custom_fields']['custom_field_64765']['value']['label']
                            project.PLUMBING_STATUS = plumbing_stage
                        # else:
                        #    .PLUMBING_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_64766']['value'] != None:
                            attic_stage = item['custom_fields']['custom_field_64766']['value']['label']
                            project.ATTIC_STATUS = attic_stage
                        # else:
                        #    .ATTIC_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_64767']['value'] != None:
                            rtv_stage = item['custom_fields']['custom_field_64767']['value']['label']
                            project.RTV_STATUS = rtv_stage
                        # else:
                        #    .RTV_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_64768']['value'] != None:
                            mev_stage = item['custom_fields']['custom_field_64768']['value']['label']
                            project.MEV_STATUS = mev_stage
                        # else:
                        #    .MEV_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_64769']['value'] != None:
                            draughtproofing_stage = item['custom_fields']['custom_field_64769']['value']['label']
                            project.DRAUGHTPROOFING_STATUS = draughtproofing_stage
                        # else:
                        #    .DRAUGHTPROOFING_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_64771']['value'] != None:
                            att_stage = item['custom_fields']['custom_field_64771']['value']['label']
                            project.AIRTIGHTNESS_STATUS = att_stage
                        # else:
                        #    .ATT_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_66383']['value'] != None:
                            solar_stage = item['custom_fields']['custom_field_66383']['value']['label']
                            project.SOLAR_STATUS = solar_stage
                        # else:
                        #    .SOLAR_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_67248']['value'] != None:
                            floor_stage = item['custom_fields']['custom_field_67248']['value']['label']
                            project.FLOOR_STATUS = floor_stage
                        # else:
                        #    .FLOOR_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_67249']['value'] != None:
                            mvhr_stage = item['custom_fields']['custom_field_67249']['value']['label']
                            project.MVHR_STATUS = mvhr_stage
                        # else:
                        #    .MVHR_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_67250']['value'] != None:
                            dcv_stage = item['custom_fields']['custom_field_67250']['value']['label']
                            project.DCV_STATUS = dcv_stage
                        # else:
                        #    .DCV_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_82331']['value'] != None:
                            co_invoices_stage = item['custom_fields']['custom_field_82331']['value']['label']
                            project.CO_INVOICES_STATUS = co_invoices_stage
                        # else:
                        #    .co-invoices_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_82333']['value'] != None:
                            final_ber_stage = item['custom_fields']['custom_field_82333']['value']['label']
                            project.FINAL_BER_STATUS = final_ber_stage
                        # else:
                        #    .FINAL_BER_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_562949953936775']['value'] != None:
                            dow_stage = item['custom_fields']['custom_field_562949953936775']['value']['label']
                            project.DOW_STATUS = dow_stage
                        # else:
                        #    .DOW_STATUS = 'N/A'
                        if item['custom_fields']['custom_field_67252']['value'] != None:
                            qc_stage = item['custom_fields']['custom_field_67252']['value']['label']
                            project.QC_STATUS = qc_stage
                        # else:
                        #    .QC_STATUS = 'N/A'

                        project_manager = get_project_manager_for_project(project_id, company_id)
                        project.Project_Manager = project_manager
                        project.save()
                        # print(project_id)
        else:
            boolean = False
            break
        page+=1        
    print('Projects updated successfully')
    return


# get_all_active_OSS_and_NISEP_projects()

# QC_Mid_Final = 'custom_field_67252'
# EWI = custom_field_64757
# CWI = custom_field_64762
# IWI = custom_field_64763
# Window = custom_field_64764
# Plumbing = custom_field_64765
# Attic = custom_field_64766
# RTV = custom_field_64767
# MEV = custom_field_64768
# Draughtproofing = custom_field_64769
# ATT = custom_field_64771
# Solar = custom_field_66383
# Floor = custom_field_67248
# MVHR = custom_field_67249
# DCV = custom_field_67250
# QC_Mid_Final = custom_field_67252
# C/O Invoices = custom_field_82331
# Final BER = custom_field_82333
# DOW = custom_field_562949953936775

# Function for getting the statuses of the custom fields lov entries from Procore and save to DB   
def get_lov_entries_statuses():
    '''
    Function gets statuses for the lov entries custom fields with status id's and saves them to DB
    '''
    DICT_OF_CUSTOM_FILEDS = {
        'QC_Mid_final' : 67252,
        'EWI' : 64757,
        'CWI' : 64762,
        'IWI' : 64763,
        'Window' : 64764,
        'Plumbing' : 64765,
        'Attic' : 64766,
        'RTV' : 64767,
        'MEV' : 64768,
        'Draughtproofing' : 64769,
        'ATT' : 64771,
        'Solar' : 66383,
        'Floor' : 67248,
        'MVHR' : 67249,
        'DCV' : 67250,
        'C/O Invoices' : 82331,
        'Final_BER' : 82333,
        'DOW' : 562949953936775
    }
    company_id = get_company_id()
    for val in DICT_OF_CUSTOM_FILEDS.values():
        access_token = get_acess_token()
        url = BASE_URL + '/rest/v1.0/custom_field_definitions/'+str(val)+'/custom_field_lov_entries'
        headers = {
            "Authorization": "Bearer " + access_token,
            'Procore-Company-Id': str(company_id)
        }
        payload = {
            'company_id' : str(company_id), 
        }
        response = requests.get(url, headers=headers, json=payload)    
        response_json = response.json()
        # print(response_json)   
        if val == 67252:
            for item in response_json:
                qc_stage = QcStages()
                qc_stage.qc_stage_id = item['id']
                qc_stage.qc_stage_name = item['label']
                qc_stage.save()
        elif val == 64757:	
            for item in response_json:
                ewi_stage = EwiStatusID()
                ewi_stage.ewi_status_id = item['id']
                ewi_stage.ewi_status_name = item['label']
                ewi_stage.save()
        elif val == 64762:
            for item in response_json:
                cwi_stage = CwiStatusID()
                cwi_stage.cwi_status_id = item['id']
                cwi_stage.cwi_status_name = item['label']
                cwi_stage.save()
        elif val == 64763:
            for item in response_json:
                iwi_stage = IwiStatusID()
                iwi_stage.iwi_status_id = item['id']
                iwi_stage.iwi_status_name = item['label']
                iwi_stage.save()
        elif val == 64764:
            for item in response_json:
                window_stage = WindowStatusID()
                window_stage.window_status_id = item['id']
                window_stage.window_status_name = item['label']
                window_stage.save()
        elif val == 64765:
            for item in response_json:
                plumbing_stage = PlumbingStatusID()
                plumbing_stage.plumbing_status_id = item['id']
                plumbing_stage.plumbing_status_name = item['label']
                plumbing_stage.save()
        elif val == 64766:
            for item in response_json:
                attic_stage = AtticStatusID()
                attic_stage.attic_status_id = item['id']
                attic_stage.attic_status_name = item['label']
                attic_stage.save()
        elif val == 64767:
            for item in response_json:
                rtv_stage = RtvStatusID()
                rtv_stage.rtv_status_id = item['id']
                rtv_stage.rtv_status_name = item['label']
                rtv_stage.save()
        elif val == 64768:
            for item in response_json:
                mev_stage = MevStatusID()
                mev_stage.mev_status_id = item['id']
                mev_stage.mev_status_name = item['label']
                mev_stage.save()
        elif val == 64769:
            for item in response_json:
                draughtproofing_stage = DraughtProofingStatusID()
                draughtproofing_stage.draught_proofing_status_id = item['id']
                draughtproofing_stage.draught_proofing_status_name = item['label']
                draughtproofing_stage.save()
        elif val == 64771:
            for item in response_json:
                att_stage = AirtightnessStatusID()
                att_stage.airtightness_status_id = item['id']
                att_stage.airtightness_status_name = item['label']
                att_stage.save()
        elif val == 66383:
            for item in response_json:
                solar_stage = SolarStatusID()
                solar_stage.solar_status_id = item['id']
                solar_stage.solar_status_name = item['label']
                solar_stage.save()
        elif val == 67248:
            for item in response_json:
                floor_stage = FloorStatusID()
                floor_stage.floor_status_id = item['id']
                floor_stage.floor_status_name = item['label']
                floor_stage.save()
        elif val == 67249:
            for item in response_json:
                mvhr_stage = MVHRStatusID()
                mvhr_stage.mvhr_status_id = item['id']
                mvhr_stage.mvhr_status_name = item['label']
                mvhr_stage.save()
        elif val == 67250:
            for item in response_json:
                dcv_stage = DCVStatusID() 
                dcv_stage.dcv_status_id = item['id']
                dcv_stage.dcv_status_name = item['label']
                dcv_stage.save()
        elif val == 82331:
            for item in response_json:
                co_invoices_stage = CoInvoicesStatusID()
                co_invoices_stage.co_invoices_status_id = item['id']
                co_invoices_stage.co_invoices_status_name = item['label']
                co_invoices_stage.save()
        elif val == 82333:
            for item in response_json:
                final_ber_stage = FinalBerStatusID()
                final_ber_stage.final_ber_status_id = item['id']
                final_ber_stage.final_ber_status_name = item['label']
                final_ber_stage.save()
        elif val == 562949953936775:
            for item in response_json:
                dow_stage = DowStatusID()
                dow_stage.dow_status_id = item['id']
                dow_stage.dow_status_name = item['label']
                dow_stage.save()
    return

#Run once a day: function saves all the company level OSS & NISEP stages in the DB wit their id
def save_company_stages():
    access_token = get_acess_token()
    company_id = str(get_company_id())
    url = BASE_URL + "/rest/v1.0/companies/"+company_id+"/project_stages"
    headers = {
        "Authorization": "Bearer " + access_token,
        'Procore-Company-Id': company_id
    }
    response = requests.get(url, headers=headers)
    response_json = response.json()
    for val in response_json:
        # print(val)
        if val['name'][0] == 'O' or val['name'][0] == 'N':
            project_object = CompanyProjectStages()
            project_object.project_stage_id = val['id']
            project_object.project_stage_name = val['name']
            project_object.save()
    return 

# get_lov_entries_statuses()
# save_company_stages()

def clean_db_for_fresh_projects():
    a = ProjectDetails.objects.all()
    a.delete()
    return


def clean_db_for_fresh_statuses():
    a = WindowStatusID.objects.all()
    a.delete()
    b = CwiStatusID.objects.all()
    b.delete()
    c = ProjectOffices.objects.all()
    c.delete()

    e = EwiStatusID.objects.all()
    e.delete()
    f = PlumbingStatusID.objects.all()
    f.delete()
    g = IwiStatusID.objects.all()
    g.delete()
    h = AirtightnessStatusID.objects.all()
    h.delete()
    i = RtvStatusID.objects.all()
    i.delete()
    j = MevStatusID.objects.all()
    j.delete()
    k = AtticStatusID.objects.all()
    k.delete()
    l = DraughtProofingStatusID.objects.all()
    l.delete()
    
    o = CompanyProjectStages.objects.all()
    o.delete()
    p = QcStages.objects.all()
    p.delete()
    q = ProjectManagers.objects.all()
    q.delete()
    r = SolarStatusID.objects.all()
    r.delete()
    s = FloorStatusID.objects.all()
    s.delete()
    t = MVHRStatusID.objects.all()
    t.delete()
    u = DCVStatusID.objects.all()
    u.delete()
    v = CoInvoicesStatusID.objects.all()
    v.delete()
    w = FinalBerStatusID.objects.all()
    w.delete()
    x = DowStatusID.objects.all()
    x.delete()
    return

# clean_db_for_fresh_projects()
# clean_db_for_fresh_statuses()

########################################################################################################  
# create your views here
########################################################################################################

@login_required(login_url='/login/')
def dashboard(request):
    '''
    Function to display the dashboard with all the projects and their statuses.
    '''
    if request.method == 'GET':
        projects_for_paging = ProjectDetails.objects.all().order_by('project_name')
        total_project_in_db = projects_for_paging.count()
    airtightness_options = AirtightnessStatusID.objects.all()
    attic_options = AtticStatusID.objects.all()
    cwi_options = CwiStatusID.objects.all()
    draught_proofing_options = DraughtProofingStatusID.objects.all()
    ewi_options = EwiStatusID.objects.all()
    iwi_options = IwiStatusID.objects.all()
    mev_options = MevStatusID.objects.all()
    plumbing_options = PlumbingStatusID.objects.all()
    rtv_options = RtvStatusID.objects.all()
    window_options = WindowStatusID.objects.all()
    stages_options = CompanyProjectStages.objects.all().order_by('project_stage_name')
    qc_options = QcStages.objects.all()
    pm_options = ProjectManagers.objects.all()
    dow_options = DowStatusID.objects.all()
    solar_options = SolarStatusID.objects.all()
    floor_options = FloorStatusID.objects.all()
    mvhr_options = MVHRStatusID.objects.all()
    dcv_options = DCVStatusID.objects.all()
    co_invoices_options = CoInvoicesStatusID.objects.all()
    final_ber_options = FinalBerStatusID.objects.all()
    project_offices = ProjectOffices.objects.all()

    # EWI = custom_field_64757 -
    # CWI = custom_field_64762 -
    # IWI = custom_field_64763 -
    # Window = custom_field_64764 -
    # Plumbing = custom_field_64765 -
    # Attic = custom_field_64766 
    # RTV = custom_field_64767 -
    # MEV = custom_field_64768 -
    # Draughtproofing = custom_field_64769 -
    # ATT = custom_field_64771 
    # Solar = custom_field_66383
    # Floor = custom_field_67248
    # MVHR = custom_field_67249
    # DCV = custom_field_67250
    # QC_Mid_Final = custom_field_67252 -
    # C/O Invoices = custom_field_82331
    # Final BER = custom_field_82333
    # DOW = custom_field_562949953936775

    if request.method == 'POST':
        if 'stage' in request.POST:
            split_str_list = request.POST['stage'].split("_")
            main_stage_name = split_str_list[0]
            project_id = split_str_list[1]  
            response_status_code = update_procore_project_stage(project_id, main_stage_name)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id) 
                project_object.project_stage = main_stage_name
                project_object.save()
        if 'qc_mid_final' in request.POST:
            split_str_list = request.POST['qc_mid_final'].split("_")
            qc_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_67252'
            custom_field_status_object = QcStages.objects.get(qc_stage_name=qc_stage_name)
            custom_field_status_id = custom_field_status_object.qc_stage_id
            print(custom_field_status_id)
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.QC_STATUS = qc_stage_name
                project_object.save()
        if 'window' in request.POST:
            split_str_list = request.POST['window'].split("_")
            window_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64764'
            custom_field_status_object = WindowStatusID.objects.get(window_status_name=window_stage_name)
            custom_field_status_id = custom_field_status_object.window_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.WINDOW_STATUS = window_stage_name
                project_object.save()
        if 'cwi' in request.POST:
            split_str_list = request.POST['cwi'].split("_")
            cwi_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64762'
            custom_field_status_object = CwiStatusID.objects.get(cwi_status_name=cwi_stage_name)
            custom_field_status_id = custom_field_status_object.cwi_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.CWI_STATUS = cwi_stage_name
                project_object.save()
        if 'ewi' in request.POST:
            split_str_list = request.POST['ewi'].split("_")
            ewi_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64757'
            custom_field_status_object = EwiStatusID.objects.get(ewi_status_name=ewi_stage_name)
            custom_field_status_id = custom_field_status_object.ewi_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.EWI_STATUS = ewi_stage_name
                project_object.save()
        if 'plumbing' in request.POST:
            split_str_list = request.POST['plumbing'].split("_")
            plumbing_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64765'
            custom_field_status_object = PlumbingStatusID.objects.get(plumbing_status_name=plumbing_stage_name)
            custom_field_status_id = custom_field_status_object.plumbing_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.PLUMBING_STATUS = plumbing_stage_name
                project_object.save()
        if 'iwi' in request.POST:
            split_str_list = request.POST['iwi'].split("_")
            iwi_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64763'
            custom_field_status_object = IwiStatusID.objects.get(iwi_status_name=iwi_stage_name)
            custom_field_status_id = custom_field_status_object.iwi_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.IWI_STATUS = iwi_stage_name
                project_object.save() 
        if 'rtv' in request.POST:
            split_str_list = request.POST['rtv'].split("_")
            rtv_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64767'
            custom_field_status_object = RtvStatusID.objects.get(rtv_status_name=rtv_stage_name)
            custom_field_status_id = custom_field_status_object.rtv_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.RTV_STATUS = rtv_stage_name
                project_object.save()
        if 'mev' in request.POST:
            split_str_list = request.POST['mev'].split("_")
            mev_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64768'
            custom_field_status_object = MevStatusID.objects.get(mev_status_name=mev_stage_name)
            custom_field_status_id = custom_field_status_object.mev_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.MEV_STATUS = mev_stage_name
                project_object.save()
        if 'attic' in request.POST:
            split_str_list = request.POST['attic'].split("_")
            attic_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64766'
            custom_field_status_object = AtticStatusID.objects.get(attic_status_name=attic_stage_name)
            custom_field_status_id = custom_field_status_object.attic_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.ATTIC_STATUS = attic_stage_name
                project_object.save()
        if 'draughtproofing' in request.POST:
            split_str_list = request.POST['draughtproofing'].split("_")
            draughtproofing_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64769'
            custom_field_status_object = DraughtProofingStatusID.objects.get(draught_proofing_status_name=draughtproofing_stage_name)
            custom_field_status_id = custom_field_status_object.draught_proofing_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.DRAUGHTPROOFING_STATUS = draughtproofing_stage_name
                project_object.save()
        if 'solar' in request.POST:
            split_str_list = request.POST['solar'].split("_")
            solar_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_66383'
            custom_field_status_object = SolarStatusID.objects.get(solar_status_name=solar_stage_name)
            custom_field_status_id = custom_field_status_object.solar_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.SOLAR_STATUS = solar_stage_name
                project_object.save()
        if 'floor' in request.POST:
            split_str_list = request.POST['floor'].split("_")
            floor_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_67248'
            custom_field_status_object = FloorStatusID.objects.get(floor_status_name=floor_stage_name)
            custom_field_status_id = custom_field_status_object.floor_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.FLOOR_STATUS = floor_stage_name
                project_object.save()
        if 'mvhr' in request.POST:
            split_str_list = request.POST['mvhr'].split("_")
            mvhr_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_67249'
            custom_field_status_object = MVHRStatusID.objects.get(mvhr_status_name=mvhr_stage_name)
            custom_field_status_id = custom_field_status_object.mvhr_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.MVHR_STATUS = mvhr_stage_name
                project_object.save()
        if 'dcv' in request.POST:
            split_str_list = request.POST['dcv'].split("_")
            dcv_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_67250'
            custom_field_status_object = DCVStatusID.objects.get(dcv_status_name=dcv_stage_name)
            custom_field_status_id = custom_field_status_object.dcv_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.DCV_STATUS = dcv_stage_name
                project_object.save()
        if 'co_invoices' in request.POST:
            split_str_list = request.POST['co_invoices'].split("_")
            co_invoices_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_82331'
            custom_field_status_object = CoInvoicesStatusID.objects.get(co_invoices_status_name=co_invoices_stage_name)
            custom_field_status_id = custom_field_status_object.co_invoices_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.CO_INVOICES_STATUS = co_invoices_stage_name
                project_object.save()
        if 'final_ber' in request.POST:
            split_str_list = request.POST['final_ber'].split("_")
            final_ber_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_82333'
            custom_field_status_object = FinalBerStatusID.objects.get(final_ber_status_name=final_ber_stage_name)
            custom_field_status_id = custom_field_status_object.final_ber_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.FINAL_BER_STATUS = final_ber_stage_name
                project_object.save()
        if 'dow' in request.POST:
            split_str_list = request.POST['dow'].split("_")
            dow_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_562949953936775'
            custom_field_status_object = DowStatusID.objects.get(dow_status_name=dow_stage_name)
            custom_field_status_id = custom_field_status_object.dow_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.DOW_STATUS = dow_stage_name
                project_object.save()
        if 'airtightness' in request.POST:
            split_str_list = request.POST['airtightness'].split("_")
            att_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64771'
            custom_field_status_object = AirtightnessStatusID.objects.get(airtightness_status_name=att_stage_name)
            custom_field_status_id = custom_field_status_object.airtightness_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.AIRTIGHTNESS_STATUS = att_stage_name
                project_object.save()
        # print(split_str_list)
        projects_for_paging = ProjectDetails.objects.all().order_by('project_name')
        total_project_in_db = projects_for_paging.count()
        
    paginator = Paginator(projects_for_paging, per_page=15, orphans=5)
    page = request.GET.get('page') # get the page number from the URL
    page_object = paginator.get_page(page)
    return render(request, 'home.html', context= {"projects" : page_object, 'airtightness_options' : airtightness_options, 'attic_options' : attic_options,
        'cwi_options' : cwi_options, 'draught_proofing_options' : draught_proofing_options, 'ewi_options' : ewi_options, 'dow_options' : dow_options,
        'iwi_options' : iwi_options, 'mev_options' : mev_options, 'pm_options' : pm_options, 'solar_options' : solar_options, 'floor_options' : floor_options,
        'mvhr_options' : mvhr_options, 'dcv_options' : dcv_options, 'co_invoices_options' : co_invoices_options,
        'final_ber_options' : final_ber_options, 'plumbing_options' : plumbing_options, 'rtv_options' : rtv_options, 'qc_options' : qc_options, 
        'window_options' : window_options, 'stages_options' : stages_options, 'project_offices' : project_offices,
        'total_project_in_db' : total_project_in_db, 'page_object':page_object})


@login_required(login_url='/login/')
def filter_projects(request):
    '''
    Function filters the projects based on the selected stage, project manager, office, QC status and CO invoices status
    '''
    global filtered_projects
    global filtered_projects_count
    global filters_input
    
    # check if request.method == 'POST' and 'filter_by_stage' in request.POST:
    if 'filter_by_stage' in request.POST:
        # print(request.POST['filter_by_pm'])
        if request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] == 'Please Select an Office' and request.POST['filter_by_qc'] == 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
            }
        elif request.POST['filter_by_stage'] == 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] == 'Please Select an Office' and request.POST['filter_by_qc'] == 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            filter_pm = request.POST['filter_by_pm']
            filtered_projects = ProjectDetails.objects.filter(Project_Manager=filter_pm).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_pm': filter_pm,
            }
        elif request.POST['filter_by_stage'] == 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] == 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            filter_office = request.POST['filter_by_office']
            filtered_projects = ProjectDetails.objects.filter(project_office=filter_office).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_office': filter_office,
            }
        elif request.POST['filter_by_stage'] == 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] == 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            filter_qc = request.POST['filter_by_qc']
            filtered_projects = ProjectDetails.objects.filter(QC_STATUS=filter_qc).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_qc': filter_qc,
            }
        elif request.POST['filter_by_stage'] == 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] == 'Please Select an Office' and request.POST['filter_by_qc'] == 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] == 'Please Select an Office' and request.POST['filter_by_qc'] == 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_pm = request.POST['filter_by_pm']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, Project_Manager=filter_pm).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_pm': filter_pm,
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] == 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_office = request.POST['filter_by_office']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, project_office=filter_office).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_office': filter_office,
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] == 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_qc = request.POST['filter_by_qc']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, QC_STATUS=filter_qc).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_qc': filter_qc,
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] == 'Please Select an Office' and request.POST['filter_by_qc'] == 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] == 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] == 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            filter_pm = request.POST['filter_by_pm']
            filter_office = request.POST['filter_by_office']
            filtered_projects = ProjectDetails.objects.filter(Project_Manager=filter_pm, project_office=filter_office).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_pm': filter_pm,
                'filter_by_office': filter_office,
            }
        elif request.POST['filter_by_stage'] == 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] == 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            filter_pm = request.POST['filter_by_pm']
            filter_qc = request.POST['filter_by_qc']
            filtered_projects = ProjectDetails.objects.filter(Project_Manager=filter_pm, QC_STATUS=filter_qc).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_pm': filter_pm,
                'filter_by_qc': filter_qc,
            }
        elif request.POST['filter_by_stage'] == 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] == 'Please Select an Office' and request.POST['filter_by_qc'] == 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_pm = request.POST['filter_by_pm']
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(Project_Manager=filter_pm, CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_pm': filter_pm,
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] == 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            filter_office = request.POST['filter_by_office']
            filter_qc = request.POST['filter_by_qc']
            filtered_projects = ProjectDetails.objects.filter(project_office=filter_office, QC_STATUS=filter_qc).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_office': filter_office,
                'filter_by_qc': filter_qc,
            }
        elif request.POST['filter_by_stage'] == 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] == 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_office = request.POST['filter_by_office']
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(project_office=filter_office, CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_office': filter_office,
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] == 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] == 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_qc = request.POST['filter_by_qc']
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(QC_STATUS=filter_qc, CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_qc': filter_qc,
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] == 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_pm = request.POST['filter_by_pm']
            filter_office = request.POST['filter_by_office']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, Project_Manager=filter_pm, project_office=filter_office).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_pm': filter_pm,
                'filter_by_office': filter_office,
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] == 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_pm = request.POST['filter_by_pm']
            filter_qc = request.POST['filter_by_qc']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, Project_Manager=filter_pm, QC_STATUS=filter_qc).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_pm': filter_pm,
                'filter_by_qc': filter_qc,
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_office = request.POST['filter_by_office']
            filter_qc = request.POST['filter_by_qc']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, project_office=filter_office, QC_STATUS=filter_qc).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_office': filter_office,
                'filter_by_qc': filter_qc,
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] == 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_qc = request.POST['filter_by_qc']
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, QC_STATUS=filter_qc, CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_qc': filter_qc,
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] == 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] == 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_pm = request.POST['filter_by_pm']
            filter_office = request.POST['filter_by_office']
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(Project_Manager=filter_pm, project_office=filter_office, CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_pm': filter_pm,
                'filter_by_office': filter_office,
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] == 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            filter_pm = request.POST['filter_by_pm']
            filter_office = request.POST['filter_by_office']
            filter_qc = request.POST['filter_by_qc']
            filtered_projects = ProjectDetails.objects.filter(Project_Manager=filter_pm, project_office=filter_office, QC_STATUS=filter_qc).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_pm': filter_pm,
                'filter_by_office': filter_office,
                'filter_by_qc': filter_qc
            }
        elif request.POST['filter_by_stage'] == 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_office = request.POST['filter_by_office']
            filter_qc = request.POST['filter_by_qc']
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(project_office=filter_office, QC_STATUS=filter_qc, CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_office': filter_office,
                'filter_by_qc': filter_qc,
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] == 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_pm = request.POST['filter_by_pm']
            filter_office = request.POST['filter_by_office']
            filter_qc = request.POST['filter_by_qc']
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(Project_Manager=filter_pm, project_office=filter_office, QC_STATUS=filter_qc, CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_pm': filter_pm,
                'filter_by_office': filter_office,
                'filter_by_qc': filter_qc,
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_office = request.POST['filter_by_office']
            filter_qc = request.POST['filter_by_qc']
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, project_office=filter_office, QC_STATUS=filter_qc, CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_office': filter_office,
                'filter_by_qc': filter_qc,
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] == 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_pm = request.POST['filter_by_pm']
            filter_qc = request.POST['filter_by_qc']
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, Project_Manager=filter_pm, QC_STATUS=filter_qc, CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_pm': filter_pm,
                'filter_by_qc': filter_qc,
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] == 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_pm = request.POST['filter_by_pm']
            filter_office = request.POST['filter_by_office']
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, Project_Manager=filter_pm, project_office=filter_office, CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_pm': filter_pm,
                'filter_by_office': filter_office,
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_pm = request.POST['filter_by_pm']
            filter_office = request.POST['filter_by_office']
            filter_qc = request.POST['filter_by_qc']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, Project_Manager=filter_pm, project_office=filter_office, QC_STATUS=filter_qc).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_pm': filter_pm,
                'filter_by_office': filter_office,
                'filter_by_qc': filter_qc
            }
            filter_stage = request.POST['filter_by_stage']
            filter_pm = request.POST['filter_by_pm']
            filter_office = request.POST['filter_by_office']
            filter_qc = request.POST['filter_by_qc']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, Project_Manager=filter_pm, project_office=filter_office, QC_STATUS=filter_qc).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_pm': filter_pm,
                'filter_by_office': filter_office,
                'filter_by_qc': filter_qc,
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] == 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_pm = request.POST['filter_by_pm']
            filter_office = request.POST['filter_by_office']
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, Project_Manager=filter_pm, project_office=filter_office, CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_pm': filter_pm,
                'filter_by_office': filter_office,
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_office = request.POST['filter_by_office']
            filter_qc = request.POST['filter_by_qc']
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, project_office=filter_office, QC_STATUS=filter_qc, CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_office': filter_office,
                'filter_by_qc': filter_qc,
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] == 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_pm = request.POST['filter_by_pm']
            filter_qc = request.POST['filter_by_qc']
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, Project_Manager=filter_pm, QC_STATUS=filter_qc, CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_pm': filter_pm,
                'filter_by_qc': filter_qc,
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] != 'Please Select a Stage' and request.POST['filter_by_pm'] != 'Please Select a Project Manager' and request.POST['filter_by_office'] != 'Please Select an Office' and request.POST['filter_by_qc'] != 'Please Select a Stage' and request.POST['filter_by_co_invoices'] != 'Please Select a Stage':
            filter_stage = request.POST['filter_by_stage']
            filter_pm = request.POST['filter_by_pm']
            filter_office = request.POST['filter_by_office']
            filter_qc = request.POST['filter_by_qc']
            filter_co = request.POST['filter_by_co_invoices']
            filtered_projects = ProjectDetails.objects.filter(project_stage=filter_stage, Project_Manager=filter_pm, project_office=filter_office, QC_STATUS=filter_qc, CO_INVOICES_STATUS=filter_co).values().order_by('project_name')
            filtered_projects_count = filtered_projects.count()
            filters_input = {
                'filter_by_stage': filter_stage,
                'filter_by_pm': filter_pm,
                'filter_by_office': filter_office,
                'filter_by_qc': filter_qc,
                'filter_by_co_invoices': filter_co
            }
        elif request.POST['filter_by_stage'] == 'Please Select a Stage' and request.POST['filter_by_pm'] == 'Please Select a Project Manager' and request.POST['filter_by_office'] == 'Please Select an Office' and request.POST['filter_by_qc'] == 'Please Select a Stage' and request.POST['filter_by_co_invoices'] == 'Please Select a Stage':
            return redirect('dashboard')
        else:
            return redirect('dashboard')
    
    airtightness_options = AirtightnessStatusID.objects.all()
    attic_options = AtticStatusID.objects.all()
    cwi_options = CwiStatusID.objects.all()
    draught_proofing_options = DraughtProofingStatusID.objects.all()
    ewi_options = EwiStatusID.objects.all()
    iwi_options = IwiStatusID.objects.all()
    mev_options = MevStatusID.objects.all()
    oil_boiler_options = OilBoilerStatusID.objects.all()
    oil_tank_options = OilTankStatusID.objects.all()
    plumbing_options = PlumbingStatusID.objects.all()
    rtv_options = RtvStatusID.objects.all()
    window_options = WindowStatusID.objects.all()
    stages_options = CompanyProjectStages.objects.all().order_by('project_stage_name')
    qc_options = QcStages.objects.all()
    pm_options = ProjectManagers.objects.all()
    dow_options = DowStatusID.objects.all()
    solar_options = SolarStatusID.objects.all()
    floor_options = FloorStatusID.objects.all()
    mvhr_options = MVHRStatusID.objects.all()
    dcv_options = DCVStatusID.objects.all()
    co_invoices_options = CoInvoicesStatusID.objects.all()
    final_ber_options = FinalBerStatusID.objects.all()
    project_offices = ProjectOffices.objects.all()

    if request.method == 'POST' and 'filter_by_stage' not in request.POST:
        if 'stage' in request.POST:
            split_str_list = request.POST['stage'].split("_")
            main_stage_name = split_str_list[0]
            project_id = split_str_list[1]  
            response_status_code = update_procore_project_stage(project_id, main_stage_name)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.project_stage = main_stage_name
                project_object.save()
        if 'qc_mid_final' in request.POST:
            split_str_list = request.POST['qc_mid_final'].split("_")
            qc_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_67252'
            custom_field_status_object = QcStages.objects.get(qc_stage_name=qc_stage_name)
            custom_field_status_id = custom_field_status_object.qc_stage_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.QC_STATUS = qc_stage_name
                project_object.save()
        if 'window' in request.POST:
            split_str_list = request.POST['window'].split("_")
            window_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64764'
            custom_field_status_object = WindowStatusID.objects.get(window_status_name=window_stage_name)
            custom_field_status_id = custom_field_status_object.window_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.WINDOW_STATUS = window_stage_name
                project_object.save()
        if 'cwi' in request.POST:
            split_str_list = request.POST['cwi'].split("_")
            cwi_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64762'
            custom_field_status_object = CwiStatusID.objects.get(cwi_status_name=cwi_stage_name)
            custom_field_status_id = custom_field_status_object.cwi_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.CWI_STATUS = cwi_stage_name
                project_object.save()
        if 'ewi' in request.POST:
            split_str_list = request.POST['ewi'].split("_")
            ewi_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64757'
            custom_field_status_object = EwiStatusID.objects.get(ewi_status_name=ewi_stage_name)
            custom_field_status_id = custom_field_status_object.ewi_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.EWI_STATUS = ewi_stage_name
                project_object.save()
        if 'plumbing' in request.POST:
            split_str_list = request.POST['plumbing'].split("_")
            plumbing_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64765'
            custom_field_status_object = PlumbingStatusID.objects.get(plumbing_status_name=plumbing_stage_name)
            custom_field_status_id = custom_field_status_object.plumbing_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.PLUMBING_STATUS = plumbing_stage_name
                project_object.save()
        if 'iwi' in request.POST:
            split_str_list = request.POST['iwi'].split("_")
            iwi_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64763'
            custom_field_status_object = IwiStatusID.objects.get(iwi_status_name=iwi_stage_name)
            custom_field_status_id = custom_field_status_object.iwi_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.IWI_STATUS = iwi_stage_name
                project_object.save() 
        if 'rtv' in request.POST:
            split_str_list = request.POST['rtv'].split("_")
            rtv_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64767'
            custom_field_status_object = RtvStatusID.objects.get(rtv_status_name=rtv_stage_name)
            custom_field_status_id = custom_field_status_object.rtv_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.RTV_STATUS = rtv_stage_name
                project_object.save()
        if 'mev' in request.POST:
            split_str_list = request.POST['mev'].split("_")
            mev_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64768'
            custom_field_status_object = MevStatusID.objects.get(mev_status_name=mev_stage_name)
            custom_field_status_id = custom_field_status_object.mev_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.MEV_STATUS = mev_stage_name
                project_object.save()
        if 'attic' in request.POST:
            split_str_list = request.POST['attic'].split("_")
            attic_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64766'
            custom_field_status_object = AtticStatusID.objects.get(attic_status_name=attic_stage_name)
            custom_field_status_id = custom_field_status_object.attic_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.ATTIC_STATUS = attic_stage_name
                project_object.save()
        if 'draughtproofing' in request.POST:
            split_str_list = request.POST['draughtproofing'].split("_")
            draughtproofing_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64769'
            custom_field_status_object = DraughtProofingStatusID.objects.get(draught_proofing_status_name=draughtproofing_stage_name)
            custom_field_status_id = custom_field_status_object.draught_proofing_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.DRAUGHTPROOFING_STATUS = draughtproofing_stage_name
                project_object.save()
        if 'solar' in request.POST:
            split_str_list = request.POST['solar'].split("_")
            solar_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_66383'
            custom_field_status_object = SolarStatusID.objects.get(solar_status_name=solar_stage_name)
            custom_field_status_id = custom_field_status_object.solar_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.SOLAR_STATUS = solar_stage_name
                project_object.save()
        if 'floor' in request.POST:
            split_str_list = request.POST['floor'].split("_")
            floor_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_67248'
            custom_field_status_object = FloorStatusID.objects.get(floor_status_name=floor_stage_name)
            custom_field_status_id = custom_field_status_object.floor_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.FLOOR_STATUS = floor_stage_name
                project_object.save()
        if 'mvhr' in request.POST:
            split_str_list = request.POST['mvhr'].split("_")
            mvhr_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_67249'
            custom_field_status_object = MVHRStatusID.objects.get(mvhr_status_name=mvhr_stage_name)
            custom_field_status_id = custom_field_status_object.mvhr_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.MVHR_STATUS = mvhr_stage_name
                project_object.save()
        if 'dcv' in request.POST:
            split_str_list = request.POST['dcv'].split("_")
            dcv_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_67250'
            custom_field_status_object = DCVStatusID.objects.get(dcv_status_name=dcv_stage_name)
            custom_field_status_id = custom_field_status_object.dcv_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.DCV_STATUS = dcv_stage_name
                project_object.save()
        if 'co_invoices' in request.POST:
            split_str_list = request.POST['co_invoices'].split("_")
            co_invoices_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_82331'
            custom_field_status_object = CoInvoicesStatusID.objects.get(co_invoices_status_name=co_invoices_stage_name)
            custom_field_status_id = custom_field_status_object.co_invoices_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.CO_INVOICES_STATUS = co_invoices_stage_name
                project_object.save()
        if 'final_ber' in request.POST:
            split_str_list = request.POST['final_ber'].split("_")
            final_ber_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_82333'
            custom_field_status_object = FinalBerStatusID.objects.get(final_ber_status_name=final_ber_stage_name)
            custom_field_status_id = custom_field_status_object.final_ber_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.FINAL_BER_STATUS = final_ber_stage_name
                project_object.save()
        if 'dow' in request.POST:
            split_str_list = request.POST['dow'].split("_")
            dow_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_562949953936775'
            custom_field_status_object = DowStatusID.objects.get(dow_status_name=dow_stage_name)
            custom_field_status_id = custom_field_status_object.dow_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.DOW_STATUS = dow_stage_name
                project_object.save()
        if 'airtightness' in request.POST:
            split_str_list = request.POST['airtightness'].split("_")
            att_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64771'
            custom_field_status_object = AirtightnessStatusID.objects.get(airtightness_status_name=att_stage_name)
            custom_field_status_id = custom_field_status_object.airtightness_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.AIRTIGHTNESS_STATUS = att_stage_name
                project_object.save()
    
    paginator = Paginator(filtered_projects, per_page=15, orphans=5)
    page = request.GET.get('page') # get the page number from the URL
    page_object = paginator.get_page(page)
    return render(request, 'filtered_projects.html', context= {"projects" : page_object, 'airtightness_options' : airtightness_options, 'attic_options' : attic_options,
        'cwi_options' : cwi_options, 'draught_proofing_options' : draught_proofing_options, 'ewi_options' : ewi_options,
        'dow_options' : dow_options, 'solar_options' : solar_options, 'floor_options' : floor_options, 'filters_input' : filters_input,
        'mvhr_options' : mvhr_options, 'dcv_options' : dcv_options, 'co_invoices_options' : co_invoices_options, 'final_ber_options' : final_ber_options,
        'iwi_options' : iwi_options, 'mev_options' : mev_options, 'oil_boiler_options' : oil_boiler_options, 'pm_options' : pm_options,
        'oil_tank_options' : oil_tank_options, 'plumbing_options' : plumbing_options, 'rtv_options' : rtv_options, 'qc_options' : qc_options,
        'window_options' : window_options, 'stages_options' : stages_options, 'project_offices' : project_offices,
        'total_project_in_db' : filtered_projects_count, 'page_object' : page_object})

   
@login_required(login_url='/login/')
def search_projects(request):
    if 'search-input' in request.POST:
        search_input = request.POST['search-input'] 
        request.session['search-input'] = search_input
    elif request.session.get('search-input'):
        search_input = request.session.get('search-input')
    else:
        redirect(dashboard)
    # print(search_input)
    searched_projects = ProjectDetails.objects.filter(project_name__icontains=str(search_input)).values().order_by('project_name')
    search_projects_count = searched_projects.count()
    airtightness_options = AirtightnessStatusID.objects.all()
    attic_options = AtticStatusID.objects.all()
    cwi_options = CwiStatusID.objects.all()
    draught_proofing_options = DraughtProofingStatusID.objects.all()
    ewi_options = EwiStatusID.objects.all()
    iwi_options = IwiStatusID.objects.all()
    mev_options = MevStatusID.objects.all()
    plumbing_options = PlumbingStatusID.objects.all()
    rtv_options = RtvStatusID.objects.all()
    window_options = WindowStatusID.objects.all()
    stages_options = CompanyProjectStages.objects.all().order_by('project_stage_name')
    qc_options = QcStages.objects.all()
    pm_options = ProjectManagers.objects.all()
    dow_options = DowStatusID.objects.all()
    solar_options = SolarStatusID.objects.all()
    floor_options = FloorStatusID.objects.all()
    mvhr_options = MVHRStatusID.objects.all()
    dcv_options = DCVStatusID.objects.all()
    co_invoices_options = CoInvoicesStatusID.objects.all()
    final_ber_options = FinalBerStatusID.objects.all()
    project_offices = ProjectOffices.objects.all()

    if request.method == 'POST':
        if 'stage' in request.POST:
            split_str_list = request.POST['stage'].split("_")
            main_stage_name = split_str_list[0]
            project_id = split_str_list[1]  
            response_status_code = update_procore_project_stage(project_id, main_stage_name)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.project_stage = main_stage_name
                project_object.save()
        if 'qc_mid_final' in request.POST:
            split_str_list = request.POST['qc_mid_final'].split("_")
            qc_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_67252'
            custom_field_status_object = QcStages.objects.get(qc_stage_name=qc_stage_name)
            custom_field_status_id = custom_field_status_object.qc_stage_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.QC_STATUS = qc_stage_name
                project_object.save()
        if 'window' in request.POST:
            split_str_list = request.POST['window'].split("_")
            window_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64764'
            custom_field_status_object = WindowStatusID.objects.get(window_status_name=window_stage_name)
            custom_field_status_id = custom_field_status_object.window_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.WINDOW_STATUS = window_stage_name
                project_object.save()
        if 'cwi' in request.POST:
            split_str_list = request.POST['cwi'].split("_")
            cwi_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64762'
            custom_field_status_object = CwiStatusID.objects.get(cwi_status_name=cwi_stage_name)
            custom_field_status_id = custom_field_status_object.cwi_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.CWI_STATUS = cwi_stage_name
                project_object.save()
        if 'ewi' in request.POST:
            split_str_list = request.POST['ewi'].split("_")
            ewi_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64757'
            custom_field_status_object = EwiStatusID.objects.get(ewi_status_name=ewi_stage_name)
            custom_field_status_id = custom_field_status_object.ewi_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.EWI_STATUS = ewi_stage_name
                project_object.save()
        if 'plumbing' in request.POST:
            split_str_list = request.POST['plumbing'].split("_")
            plumbing_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64765'
            custom_field_status_object = PlumbingStatusID.objects.get(plumbing_status_name=plumbing_stage_name)
            custom_field_status_id = custom_field_status_object.plumbing_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.PLUMBING_STATUS = plumbing_stage_name
                project_object.save()
        if 'iwi' in request.POST:
            split_str_list = request.POST['iwi'].split("_")
            iwi_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64763'
            custom_field_status_object = IwiStatusID.objects.get(iwi_status_name=iwi_stage_name)
            custom_field_status_id = custom_field_status_object.iwi_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.IWI_STATUS = iwi_stage_name
                project_object.save() 
        if 'rtv' in request.POST:
            split_str_list = request.POST['rtv'].split("_")
            rtv_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64767'
            custom_field_status_object = RtvStatusID.objects.get(rtv_status_name=rtv_stage_name)
            custom_field_status_id = custom_field_status_object.rtv_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.RTV_STATUS = rtv_stage_name
                project_object.save()
        if 'mev' in request.POST:
            split_str_list = request.POST['mev'].split("_")
            mev_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64768'
            custom_field_status_object = MevStatusID.objects.get(mev_status_name=mev_stage_name)
            custom_field_status_id = custom_field_status_object.mev_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.MEV_STATUS = mev_stage_name
                project_object.save()
        if 'attic' in request.POST:
            split_str_list = request.POST['attic'].split("_")
            attic_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64766'
            custom_field_status_object = AtticStatusID.objects.get(attic_status_name=attic_stage_name)
            custom_field_status_id = custom_field_status_object.attic_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.ATTIC_STATUS = attic_stage_name
                project_object.save()
        if 'draughtproofing' in request.POST:
            split_str_list = request.POST['draughtproofing'].split("_")
            draughtproofing_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64769'
            custom_field_status_object = DraughtProofingStatusID.objects.get(draught_proofing_status_name=draughtproofing_stage_name)
            custom_field_status_id = custom_field_status_object.draught_proofing_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.DRAUGHTPROOFING_STATUS = draughtproofing_stage_name
                project_object.save()
        if 'solar' in request.POST:
            split_str_list = request.POST['solar'].split("_")
            solar_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_66383'
            custom_field_status_object = SolarStatusID.objects.get(solar_status_name=solar_stage_name)
            custom_field_status_id = custom_field_status_object.solar_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.SOLAR_STATUS = solar_stage_name
                project_object.save()
        if 'floor' in request.POST:
            split_str_list = request.POST['floor'].split("_")
            floor_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_67248'
            custom_field_status_object = FloorStatusID.objects.get(floor_status_name=floor_stage_name)
            custom_field_status_id = custom_field_status_object.floor_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.FLOOR_STATUS = floor_stage_name
                project_object.save()
        if 'mvhr' in request.POST:
            split_str_list = request.POST['mvhr'].split("_")
            mvhr_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_67249'
            custom_field_status_object = MVHRStatusID.objects.get(mvhr_status_name=mvhr_stage_name)
            custom_field_status_id = custom_field_status_object.mvhr_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.MVHR_STATUS = mvhr_stage_name
                project_object.save()
        if 'dcv' in request.POST:
            split_str_list = request.POST['dcv'].split("_")
            dcv_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_67250'
            custom_field_status_object = DCVStatusID.objects.get(dcv_status_name=dcv_stage_name)
            custom_field_status_id = custom_field_status_object.dcv_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.DCV_STATUS = dcv_stage_name
                project_object.save()
        if 'co_invoices' in request.POST:
            split_str_list = request.POST['co_invoices'].split("_")
            co_invoices_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_82331'
            custom_field_status_object = CoInvoicesStatusID.objects.get(co_invoices_status_name=co_invoices_stage_name)
            custom_field_status_id = custom_field_status_object.co_invoices_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.CO_INVOICES_STATUS = co_invoices_stage_name
                project_object.save()
        if 'final_ber' in request.POST:
            split_str_list = request.POST['final_ber'].split("_")
            final_ber_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_82333'
            custom_field_status_object = FinalBerStatusID.objects.get(final_ber_status_name=final_ber_stage_name)
            custom_field_status_id = custom_field_status_object.final_ber_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.FINAL_BER_STATUS = final_ber_stage_name
                project_object.save()
        if 'dow' in request.POST:
            split_str_list = request.POST['dow'].split("_")
            dow_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_562949953936775'
            custom_field_status_object = DowStatusID.objects.get(dow_status_name=dow_stage_name)
            custom_field_status_id = custom_field_status_object.dow_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.DOW_STATUS = dow_stage_name
                project_object.save()
        if 'airtightness' in request.POST:
            split_str_list = request.POST['airtightness'].split("_")
            att_stage_name = split_str_list[0]
            project_id = split_str_list[1]
            custom_field_id = 'custom_field_64771'
            custom_field_status_object = AirtightnessStatusID.objects.get(airtightness_status_name=att_stage_name)
            custom_field_status_id = custom_field_status_object.airtightness_status_id
            response_status_code = update_procore_project(project_id, custom_field_id, custom_field_status_id)
            if response_status_code == 200:
                project_object = ProjectDetails.objects.get(project_id=project_id)
                project_object.AIRTIGHTNESS_STATUS = att_stage_name
                project_object.save()

    paginator = Paginator(searched_projects, per_page=15, orphans=5)
    page = request.GET.get('page') # get the page number from the URL
    page_object = paginator.get_page(page)
    return render(request, 'searched_projects.html', context= {"projects" : page_object, 'airtightness_options' : airtightness_options, 'attic_options' : attic_options,
        'cwi_options' : cwi_options, 'draught_proofing_options' : draught_proofing_options, 'ewi_options' : ewi_options,
        'dow_options' : dow_options, 'solar_options' : solar_options, 'floor_options' : floor_options,
        'mvhr_options' : mvhr_options, 'dcv_options' : dcv_options, 'co_invoices_options' : co_invoices_options, 'final_ber_options' : final_ber_options,
        'iwi_options' : iwi_options, 'mev_options' : mev_options, 'pm_options' : pm_options, 'plumbing_options' : plumbing_options, 'rtv_options' : rtv_options, 'qc_options' : qc_options,
        'window_options' : window_options, 'stages_options' : stages_options, 'project_offices' : project_offices,
        'total_project_in_db' : search_projects_count, 'search_input' : search_input, 'page_object' : page_object})


@login_required(login_url='/login/')
def export_to_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="projects_excel.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'OSS & NISEP Projects'

    # Write header row
    header = ['Project ID', 'Project Name', 'Project Stage', 'QC Mid/Final', 'EWI Status', 'CWI Status', 'Window Status', 'Plumbing Status', 'IWI Status', 'RTV Status', 'MEV Status', 'Attic Status', 'Draughtproofing Status', 'Solar Status', 'Floor Status', 'MVHR Status', 'DCV Status', 'CO Invoices Status', 'Final BER Status', 'DOW Status']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    queryset = ProjectDetails.objects.all().values_list('project_number', 'project_stage', 'project_name', 'QC_STATUS', 'EWI_STATUS', 'CWI_STATUS', 'WINDOW_STATUS', 'PLUMBING_STATUS', 'IWI_STATUS', 'RTV_STATUS', 'MEV_STATUS', 'ATTIC_STATUS', 'DRAUGHTPROOFING_STATUS', 'SOLAR_STATUS', 'FLOOR_STATUS', 'MVHR_STATUS', 'DCV_STATUS', 'CO_INVOICES_STATUS', 'FINAL_BER_STATUS', 'DOW_STATUS')
    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num+1, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response

########################################################################################################
# Procore API functions for updating Procore and updating database
########################################################################################################

def update_procore_project_stage(project_id, project_stage):
    access_token = get_acess_token()
    company_id = get_company_id()
    # print(project_stage)
    project_stage_object = CompanyProjectStages.objects.get(project_stage_name=project_stage)
    project_stage_id = project_stage_object.project_stage_id
    # print(project_id, project_stage_id)
    url = BASE_URL + '/rest/v1.0/projects/'+str(project_id)
    headers = {
        "Authorization": "Bearer " + access_token,
        'Procore-Company-Id': str(company_id),
        'content-type': "application/json",
    }
    payload = {
        'company_id' : str(company_id),
        'project' : {
            'project_stage_id' : project_stage_id, 
        }
    }
    # print(payload)
    response = requests.patch(url, headers=headers, json=payload)
    response_status_code = response.status_code
    # print(response_status_code)
    return response_status_code

def update_procore_project(project_id, custom_field_id, custom_field_status_id):
    access_token = get_acess_token()
    company_id = get_company_id()
    url = BASE_URL + '/rest/v1.0/projects/'+str(project_id)
    headers = {
        "Authorization": "Bearer " + access_token,
        'Procore-Company-Id': str(company_id),
        'content-type': "application/json",
    }    
    payload = {
        'company_id' : str(company_id),
        'run_configurable_validations' : True,
        'project' : {
            custom_field_id : custom_field_status_id
        }
    }
    # print(payload)
    response = requests.patch(url, headers=headers, json=payload)
    response_status_code = response.status_code
    # print(response_status_code)
    # response_json = response.json()
    return response_status_code


#######################################################################################################
# Scheduling functions to keep DB uptodate
#######################################################################################################

scheduler = BackgroundScheduler(daemon=True, job_defaults={'max_instances': 10})
scheduler.add_job(lambda: get_all_active_OSS_and_NISEP_projects(), 'interval', minutes=90)

scheduler.add_job(lambda: clean_db_for_fresh_projects(), 'cron', hour=1)
scheduler.add_job(lambda: clean_db_for_fresh_statuses(), 'cron', hour=1)
scheduler.add_job(lambda: get_lov_entries_statuses(), 'cron', hour=1)
scheduler.add_job(lambda: save_company_stages(), 'cron', hour=1)


# scheduler.add_job(lambda : scheduler.print_jobs(),'interval',seconds=5)
scheduler.start()

#######################################################################################################
# Password Reset functionality
########################################################################################################

from django.contrib.auth.views import PasswordResetView
import logging

logger = logging.getLogger(__name__)

class DebugPasswordResetView(PasswordResetView):
    def form_valid(self, form):
        logger.debug("Sending password reset email to: %s", form.cleaned_data.get('email'))
        return super().form_valid(form)

#######################################################################################################
# The End 
#######################################################################################################


## Custom Sign Up View
from .forms import CustomUserCreationForm # Import your custom form
from django.contrib import messages

def signup_view(request):
    print(">>>>>>>>>> CUSTOM SIGNUP VIEW (accounts.views.signup_view) IS BEING CALLED! <<<<<<<<<<") # Add this line
    print(request.POST)
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            messages.success(request, f'Account created for {username}! You can now log in.')
            return redirect('login') # Redirect to your login page name
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/signup.html', {'form': form})

## Custom Password Reset View
from django.contrib.auth.views import PasswordResetView
from django.contrib.auth.forms import PasswordResetForm # Default form
from django.conf import settings # To check EMAIL_BACKEND

class CustomPasswordResetView(PasswordResetView):
    def form_valid(self, form):
        print("CustomPasswordResetView: form_valid called.")
        print(f"CustomPasswordResetView: Using EMAIL_BACKEND: {settings.EMAIL_BACKEND}")
        print(f"CustomPasswordResetView: Email to be sent to: {form.cleaned_data['email']}")
        # opts = {
        #     'use_https': self.request.is_secure(),
        #     'token_generator': self.token_generator,
        #     'from_email': self.from_email, # Or explicitly your SMTP2Go from email
        #     'email_template_name': self.email_template_name,
        #     'subject_template_name': self.subject_template_name,
        #     'request': self.request,
        #     'html_email_template_name': self.html_email_template_name,
        #     'extra_email_context': self.extra_email_context,
        # }
        # # The form.save() method is what actually sends the email.
        # # We'll try to see if it's called and if it raises an exception.
        # try:
        #     print(f"CustomPasswordResetView: Attempting to call form.save() for email: {form.cleaned_data['email']}")
        #     # If you simplified templates, ensure they are still specified or remove template_name args above
        #     # to use defaults if you reverted those changes.
        #     form.save(**opts)
        #     print("CustomPasswordResetView: form.save() completed without raising an immediate error.")
        # except Exception as e:
        #     print(f"CustomPasswordResetView: ERROR during form.save(): {e}")
        #     # You might want to re-raise the exception or handle it
        #     # For debugging, just printing it is often enough to see it in console.
        #     # raise # Uncomment to see the full traceback if it's caught silently
        return super().form_valid(form) # Call the original form_valid logic
    
# Custom Email Login View
from .forms import EmailLoginForm
from django.contrib.auth import authenticate
from django.contrib.auth import login

def email_login_view(request):
    form = EmailLoginForm(request.POST or None)
    if form.is_valid():
        email = form.cleaned_data['email']
        password = form.cleaned_data['password']
        user = authenticate(request, email=email, password=password)
        if user:
            login(request, user)
            return redirect(request.GET.get('next') or 'dashboard')  # or wherever
        else:
            messages.error(request, "Invalid email or password.")
        if user:
            print(f"🔐 Login successful for: {user.email}")
            login(request, user)
            return redirect(request.GET.get('next') or 'dashboard')
        else:
            print("❌ Authentication failed")   
    return render(request, 'registration/login.html', {'form': form})


# Custom Password Reset Confirm View
from django.contrib.auth.views import PasswordResetConfirmView

class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    def form_valid(self, form):
        print("✅ form_valid called — setting new password")
        user = form.save()  # 🛠️ This sets the new password
        print(f"✅ Password reset for user: {user.email}")
        return super().form_valid(form)
    
